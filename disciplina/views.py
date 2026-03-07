import datetime
from tempfile import NamedTemporaryFile
from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import permission_required
from django.http import HttpResponse
from openpyxl import Workbook

from escuela.models import PeriodoEscolar
from .models import FaltaDisciplinariaEstudiantil, DemeritoDeEstudiante, RedencionDeEstudiante


@permission_required("disciplina.can_change_faltadisciplinaria")
def exportar_faltas_disciplinarias_de_periodo_escolar_activo(request):
    """
    vista que exporta a un archivo de Excel las faltas disciplinarias
    de todos los estudiantes durante el período activo, ordenados por
    seccion y luego por estudiante, si un estudiante tiene varias faltas
    deben salir consecutivamente por fecha.
    """
    periodo_escolar_activo = PeriodoEscolar.objects.get(periodo_activo=True)
    faltas_de_periodo_escolar_activo = (
        FaltaDisciplinariaEstudiantil.objects.filter(
            periodo_escolar=periodo_escolar_activo
        )
        .select_related("estudiante", "estudiante__seccion__nivel_educativo", "falta")
        .order_by("estudiante__seccion", "estudiante__apellidos", "estudiante__nombre")
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Faltas de estudiantes"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 12
    ws.append(
        [
            "Estudiante",
            "Sección",
            "Tipo de Falta",
            "Falta Cometida",
            "Descripción de la falta",
            "Fecha",
        ]
    )
    for falta in faltas_de_periodo_escolar_activo:
        ws.append(
            [
                str(falta.estudiante),
                str(falta.estudiante.seccion),
                str(falta.falta.get_categoria_display()),
                str(falta.falta.descripcion),
                falta.descripcion,
                falta.fecha,
            ]
        )

    # agrego hoja de deméritos de estudiantes
    ws2 = wb.create_sheet(title="Deméritos de estudiantes")
    ws2.title = "Demeritos de estudiantes"
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 11
    ws2.column_dimensions["D"].width = 50
    ws2.column_dimensions["E"].width = 50
    ws2.column_dimensions["F"].width = 12
    ws2.append(
        [
            "Estudiante",
            "Sección",
            "Código",
            "Demérito Cometido",
            "Descripción del demérito",
            "Fecha",
        ]
    )
    demeritos_de_periodo_escolar_activo = (
        DemeritoDeEstudiante.objects.filter(
            periodo_escolar=periodo_escolar_activo
        )
        .select_related("estudiante", "estudiante__seccion__nivel_educativo", "demerito")
        .order_by("estudiante__seccion", "estudiante__apellidos", "estudiante__nombre")
    )
    for demerito in demeritos_de_periodo_escolar_activo:
        ws2.append(
            [
                str(demerito.estudiante),
                str(demerito.estudiante.seccion),
                str(demerito.demerito.codigo),
                str(demerito.demerito.descripcion),
                demerito.descripcion,
                demerito.fecha,
            ]
        )
    
    # agrego hoja de redenciones de estudiantes
    ws3 = wb.create_sheet(title="Redenciones de estudiantes")
    ws3.title = "Redenciones de estudiantes"
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 25
    ws3.column_dimensions["C"].width = 11
    ws3.column_dimensions["D"].width = 50
    ws3.column_dimensions["E"].width = 50
    ws3.column_dimensions["F"].width = 12
    ws3.append(
        [
            "Estudiante",
            "Sección",
            "Código",
            "Redención Cometida",
            "Descripción de la redención",
            "Fecha",
        ]
    )
    redenciones_de_periodo_escolar_activo = (
        RedencionDeEstudiante.objects.filter(
            periodo_escolar=periodo_escolar_activo
        )
        .select_related("estudiante", "estudiante__seccion__nivel_educativo", "redencion")
        .order_by("estudiante__seccion", "estudiante__apellidos", "estudiante__nombre")
    )
    for redencion in redenciones_de_periodo_escolar_activo:
        ws3.append(
            [
                str(redencion.estudiante),
                str(redencion.estudiante.seccion),
                str(redencion.redencion.codigo),
                str(redencion.redencion.descripcion),
                redencion.descripcion,
                redencion.fecha,
            ]
        )

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/ms-excel",
        )
        response["Content-Disposition"] = (
            f'attachment; filename=Faltas disciplinarias de estudiantes-{datetime.date.today().strftime("%Y_%m_%d")}.xlsx'
        )
    return response

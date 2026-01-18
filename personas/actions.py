import datetime, io
import openpyxl
import xlsxwriter
from tempfile import NamedTemporaryFile

from django.http import HttpResponse
from openpyxl import Workbook


def exportar_todos_los_datos_a_excel(self, request, queryset):
    """
    Acción de Django Admin que exporta a Excel todos los datos del modelo
    Estudiante, de todos los estudiantes en el queryset, haciendo uso de
    la librería openpyxl ver: openpyxl.readthedocs.io
    """

    import datetime
    
    print("Antes de queryset: ", datetime.datetime.now())
    queryset = queryset.select_related(
        "escuela_previa",
        "seccion",
        "municipio_de_nacimiento",
        "municipio_de_residencia",
        "responsable",
    ).prefetch_related("secciones", "estudiantes_en_la_misma_casa")

    print("Después de queryset y al crear archivo: ", datetime.datetime.now())

    meta = self.model._meta
    field_names = [field.name for field in meta.fields]

    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes - Datos completos"

    ws.append(field_names)

    for obj in queryset:
        ws.append([str(getattr(obj, field)) for field in field_names])
    
    print("Después de crear el archivo: ", datetime.datetime.now())


    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/ms-excel",
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename=Datos_completos_de_estudiantes-{datetime.datetime.now().strftime("%Y_%m_%d_%H-%M")}.xlsx'
        return response


def exportar_datos_de_contacto_a_excel(self, request, queryset):
    """
    Acción de Django Admin que exporta a Excel datos de contacto del modelo
    Estudiante, de todos los estudiantes en el queryset, haciendo uso de
    la librería openpyxl ver: openpyxl.readthedocs.io
    """
    queryset = queryset.select_related("responsable")
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes - contactos"
    ws.append(
        [
            "Apellidos",
            "Nombres",
            "NIE",
            "Sexo",
            "Edad",
            "Sección",
            "Teléfono 1",
            "Teléfono 2",
            "Correo Electrónico",
            "Nombre de responsable",
            "Relación",
            "DUI",
            "Teléfono 1",
            "Teléfono 2",
            "Correo Electrónico",
        ]
    )
    ws.column_dimensions["A"].width = 17.60
    ws.column_dimensions["B"].width = 17.60
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 4.50
    ws.column_dimensions["E"].width = 4.50
    ws.column_dimensions["F"].width = 31.00
    ws.column_dimensions["G"].width = 9.4
    ws.column_dimensions["H"].width = 9.4
    ws.column_dimensions["I"].width = 29.0
    ws.column_dimensions["J"].width = 31.0
    ws.column_dimensions["K"].width = 7.25
    ws.column_dimensions["L"].width = 10.30
    ws.column_dimensions["M"].width = 9.4
    ws.column_dimensions["N"].width = 9.4
    ws.column_dimensions["O"].width = 29.0
    for obj in queryset:
        responsable = obj.responsable
        if responsable is None:
            ws.append(
                [
                    obj.apellidos,
                    obj.nombre,
                    obj.nie,
                    obj.sexo,
                    obj.edad,
                    obj.seccion.__str__().title(),
                    obj.telefono_1,
                    obj.telefono_2,
                    obj.correo_electronico,
                ]
            )
        else:
            ws.append(
                [
                    obj.apellidos,
                    obj.nombre,
                    obj.nie,
                    obj.sexo,
                    obj.edad,
                    obj.seccion.__str__().title(),
                    obj.telefono_1,
                    obj.telefono_2,
                    obj.correo_electronico,
                    f"{responsable.nombre} {responsable.apellidos}",
                    obj.get_relacion_de_responsable_display(),
                    responsable.dui,
                    responsable.telefono_1,
                    responsable.telefono_2,
                    responsable.correo_electronico,
                ]
            )

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/ms-excel",
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename=DatosDeContactoDeEstudiantes-{datetime.datetime.now().strftime("%Y_%m_%d-%H%M")}.xlsx'
        return response


def exportar_datos_basicos_a_excel(self, request, queryset):
    """
    Acción de Django Admin que exporta a Excel los datos básicos del modelo
    Estudiante, de todos los estudiantes en el queryset, haciendo uso de
    la librería openpyxl ver: openpyxl.readthedocs.io
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes - Datos básicos"
    ws.append(
        [
            "Apellidos",
            "Nombres",
            "NIE",
            "Sexo",
            "Edad",
            "Sección",
            "Teléfono 1",
            "Correo Electrónico",
        ]
    )
    ws.column_dimensions["A"].width = 17.50
    ws.column_dimensions["B"].width = 19.20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 4.50
    ws.column_dimensions["E"].width = 4.50
    ws.column_dimensions["F"].width = 31.00
    ws.column_dimensions["G"].width = 9.50
    ws.column_dimensions["H"].width = 40.00
    for obj in queryset:
        ws.append(
            [
                obj.apellidos,
                obj.nombre,
                obj.nie,
                obj.sexo,
                obj.edad,
                obj.seccion.__str__().title(),
                obj.telefono_1,
                obj.correo_electronico,
            ]
        )

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/ms-excel",
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename=ListaDeEstudiantes-{datetime.datetime.now().strftime("%Y_%m_%d-%H%M")}.xlsx'
        return response


def exportar_a_excel_datos_completos_de_responsables(self, request, queryset):
    """
    Acción que toma un queryset de Responsable y exporta
    a excel las columnas: Nombre, Apellidos, DUI, Sexo, Fecha de nacimiento,
    teléfono 1 y 2, correo electrónico, Situación laboral, Dirección
    y observaciones, haciendo uso de
    la librería openpyxl ver: openpyxl.readthedocs.io
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Responsables - Estudiantes"
    ws.append(
        [
            "Apellidos",
            "Nombre",
            "DUI",
            "Sexo",
            "Nacimiento",
            "Teléfono 1",
            "Teléfono 2",
            "Correo Electrónico",
            "Situación Laboral",
            "Dirección",
            "Observaciones",
        ]
    )
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 65
    ws.column_dimensions["K"].width = 65
    for obj in queryset:
        ws.append(
            [
                obj.apellidos,
                obj.nombre,
                obj.dui,
                obj.get_sexo_display(),
                obj.fecha_de_nacimiento,
                obj.telefono_1,
                obj.telefono_2,
                obj.correo_electronico,
                obj.get_situacion_laboral_display(),
                obj.direccion_de_residencia,
                obj.observaciones,
            ]
        )

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/ms-excel",
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename=ListaDeResponsables-{datetime.datetime.now().strftime("%Y_%m_%d-%H%M")}.xlsx'
        return response


def exportar_a_excel_estudiantes_y_responsables_por_familia_y_seccion(
    self, request, queryset
):
    """
    Acción que toma un queryset de Estudiantes y exporta
    a excel las columnas: Responsable, DUI de responsable,
    estudiante y sección del estudiante, ordenados en primer
    lugar por sección, pero con familias en conjunto, es decir
    el primer niño de kinder 3 es el primero, si tiene hermanos
    luego ellos, luego el segundo niño de kinder 3 y sus hermanos y así
    sucesivamente, estos hermanos ya no aparecen en su repectiva seccion
    haciemos uso de la librería openpyxl ver: openpyxl.readthedocs.io
    """

    # optimización del queryset para los datos
    # de responsable relacionados y ordenados
    # para ello pasamos el query a values
    estudiantes = (
        queryset.select_related("responsable")
        .order_by("seccion__nivel_educativo", "seccion", "apellidos", "nombre")
        .values(
            "responsable__id",
            "responsable__apellidos",
            "responsable__nombre",
            "responsable__dui",
            "apellidos",
            "nombre",
            "seccion__seccion",
            "seccion__nivel_educativo__nivel"
        )
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Responsables - Estudiantes"
    ws.append(["Responsable", "DUI de responsable", "Estudiante", "Sección"])
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 45

    # el algoritmo para encontrar otros estudiantes del mismo
    # responsable se hace a nivel de Python y no ORM
    responsables_usados = []
    count = 0
    for estudiante in estudiantes:
        if estudiante["responsable__id"] in responsables_usados:
            count += 1
        else:
            ws.append(
                [
                    f"{estudiante['responsable__apellidos']}, {estudiante['responsable__nombre']}",
                    estudiante["responsable__dui"],
                    f"{estudiante['apellidos']}, {estudiante['nombre']}",
                    f"{estudiante['seccion__nivel_educativo__nivel']} {estudiante['seccion__seccion']}"
                ]
            )
            for estudiante_b in estudiantes[count + 1 :]:
                if estudiante["responsable__id"] == estudiante_b["responsable__id"]:
                    ws.append(
                        [
                            f"{estudiante_b['responsable__apellidos']}, {estudiante_b['responsable__nombre']}",
                            estudiante_b["responsable__dui"],
                            f"{estudiante_b['apellidos']}, {estudiante_b['nombre']}",
                            f"{estudiante_b['seccion__nivel_educativo__nivel']} {estudiante_b['seccion__seccion']}"
                        ]
                    )
            responsables_usados.append(estudiante["responsable__id"])
            count += 1

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/ms-excel",
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename=ListaDeEstudiantesYResponsables-{datetime.datetime.now().strftime("%Y_%m_%d-%H%M")}.xlsx'
        return response


def exportar_a_excel_estudiantes_y_responsables_por_seccion(self, request, queryset):
    """
    Acción que toma un queryset de Estudiantes y exporta
    a excel las columnas: Responsable, DUI de responsable,
    estudiante y sección del estudiante ordenados por seccion,
    haciendo uso de la librería openpyxl ver: openpyxl.readthedocs.io
    """

    # optimización del queryset para los datos
    # de responsable relacionados y ordenados
    # para ello pasamos el query a values
    estudiantes = (
        queryset.select_related("responsable")
        .order_by("seccion__nivel_educativo", "seccion", "apellidos", "nombre")
        .values(
            "responsable__id",
            "responsable__apellidos",
            "responsable__nombre",
            "responsable__dui",
            "apellidos",
            "nombre",
            "seccion__seccion",
            "seccion__nivel_educativo__nivel",
        )
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Responsables - Estudiantes"
    ws.append(["Responsable", "DUI de responsable", "Estudiante", "Sección"])
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 45

    # utilizamos los values
    for estudiante in estudiantes:
        ws.append(
            [
                f"{estudiante['responsable__apellidos']}, {estudiante['responsable__nombre']}",
                estudiante["responsable__dui"],
                f"{estudiante['apellidos']}, {estudiante['nombre']}",
                f"{estudiante['seccion__nivel_educativo__nivel']} {estudiante['seccion__seccion']}"
            ]
        )

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/ms-excel",
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename=ListaDeEstudiantesYResponsables-{datetime.datetime.now().strftime("%Y_%m_%d-%H%M")}.xlsx'
        return response


def exportar_a_excel_lista_de_firmas_por_seccion_y_familia(self, request, queryset):
    """
    TODO: ESTA ACCIÓN NO FUNCIONA ACTUALMENTE
    Acción que toma un queryset de Estudiantes y exporta
    a excel las columnas: Responsable, DUI de responsable,
    estudiante y sección del estudiante, ordenados en primer
    lugar por sección, pero con familias en conjunto, es decir
    si le primer nivel es kinder 3, el primer nino de ese grado aparece
    seguido de sus heramnos, luego el segundo niño de kinder 3 y sus hermanos y así
    sucesivamente, estos hermanos ya no aparecen en su repectiva seccion
    haciemos uso de la librería XlsxWriter ver: xlsxwriter.readthedocs.io
    """

    estudiantes = (
        queryset.select_related("seccion__nivel_educativo")
        .order_by("seccion__nivel_educativo", "seccion", "apellidos", "nombre")
        .values(
            "responsable__id",
            "responsable__apellidos",
            "responsable__nombre",
            "responsable__dui",
            "apellidos",
            "nombre",
            "seccion__seccion",
            "seccion__id",
            "seccion__nivel_educativo__nivel",
        )
    )

    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    header_format = wb.add_format({"bold": 1, "align": "center", "valign": "vcenter"})
    list_header_format = wb.add_format(
        {"bold": 1, "border": 1, "bg_color": "#BBBBBB", "valign": "vcenter"}
    )
    list_body_format = wb.add_format(
        {
            "border": 1,
        }
    )

    responsables_usados = []
    count = 0
    seccion_actual = "0"
    counter_interno = 1
    ws = None
    for estudiante in estudiantes:
        if estudiante["responsable__id"] in responsables_usados:
            count += 1
        else:
            if estudiante["seccion__id"] != seccion_actual:
                ws = wb.add_worksheet(
                    f"{estudiante['seccion__nivel_educativo__nivel']} {estudiante['seccion__seccion']}".title()
                    .replace("Años", "")
                    .replace("Año", "")
                    .replace("Bachillerato", "")
                )
                ws.set_column(0, 0, 3)
                ws.set_column(1, 1, 36)
                ws.set_column(2, 2, 11)
                ws.set_column(3, 3, 36)
                ws.set_column(4, 4, 40)
                ws.set_column(5, 5, 20)
                ws.merge_range(
                    "A1:F1",
                    'Complejo Educativo "Dr. Humberto Romero Alvergue"',
                    header_format,
                )
                ws.merge_range(
                    "A2:F2", "Código de Infraestructura 11674", header_format
                )
                ws.merge_range("A3:F3", "Descripción de actividad")
                ws.merge_range("A4:F4", "Fecha de entrega")
                ws.merge_range("A5:F5", "Se recibe")
                ws.merge_range("A6:F6", "")
                ws.write_row(
                    "A7",
                    [
                        "Nº",
                        "Nombre de Responsable",
                        "Nº de DUI",
                        "Estudiante",
                        "Sección de estudiante",
                        "Firma",
                    ],
                    list_header_format,
                )
                counter_interno = 1
            ws.write_row(
                counter_interno + 6,
                0,
                [
                    str(counter_interno),
                    f"{estudiante['responsable__apellidos']}, {estudiante['responsable__nombre']}",
                    estudiante["responsable__dui"],
                    f"{estudiante['apellidos']}, {estudiante['nombre']}",
                    f"{estudiante['seccion__nivel_educativo__nivel']} {estudiante['seccion__seccion']}",
                    "",
                ],
                list_body_format,
            )
            ws.set_row(counter_interno + 6, 25)
            counter_interno += 1
            for estudiante_b in estudiantes[count + 1 :]:
                if estudiante["responsable__id"] == estudiante_b["responsable__id"]:
                    ws.write_row(
                        counter_interno + 6,
                        0,
                        [
                            str(counter_interno),
                            f"{estudiante_b['responsable__apellidos']}, {estudiante_b['responsable__nombre']}",
                            estudiante_b["responsable__dui"],
                            f"{estudiante_b['apellidos']}, {estudiante_b['nombre']}",
                            f"{estudiante['seccion__nivel_educativo__nivel']} {estudiante['seccion__seccion']}",
                            "",
                        ],
                        list_body_format,
                    )
                    ws.set_row(counter_interno + 6, 25)
                    counter_interno += 1
            seccion_actual = estudiante["seccion__id"]
            responsables_usados.append(estudiante["responsable__id"])
            count += 1
    wb.close()
    output.seek(0)
    filename = f"ListaDeEstudiantesYResponsablesParaFirmas-{datetime.datetime.now().strftime('%Y_%m_%d-%H%M')}.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=%s" % filename
    return response



def exportar_a_excel_estudiantes_y_responsables_por_familia_y_seccion_separadas(
    self, request, queryset
):
    """
    Acción que toma un queryset de Estudiantes y exporta
    a excel las columnas: Responsable, DUI de responsable,
    estudiante y sección del estudiante, ordenados en primer
    lugar por sección, pero con familias en conjunto, es decir
    el primer niño de kinder 3 es el primero, si tiene hermanos
    luego ellos, luego el segundo niño de kinder 3 y sus hermanos y así
    sucesivamente, estos hermanos ya no aparecen en su repectiva seccion
    haciemos uso de la librería openpyxl ver: openpyxl.readthedocs.io
    """

    # optimización del queryset para los datos
    # de responsable relacionados y ordenados
    # para ello pasamos el query a values
    estudiantes = (
        queryset.select_related("responsable")
        .order_by("seccion__nivel_educativo", "seccion", "apellidos", "nombre")
        .values(
            "responsable__id",
            "responsable__apellidos",
            "responsable__nombre",
            "responsable__dui",
            "apellidos",
            "nombre",
            "seccion__seccion",
            "seccion__nivel_educativo__nivel"
        )
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Responsables - Estudiantes"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 45

    # el algoritmo para encontrar otros estudiantes del mismo
    # responsable se hace a nivel de Python y no ORM
    responsables_usados = []
    count = 0
    seccion = []
    contador_seccion = 1
    for estudiante in estudiantes:
        if seccion != [estudiante['seccion__nivel_educativo__nivel'], estudiante['seccion__seccion']]:
            ws.append([""])
            ws.append([""])
            ws.append(["", estudiante['seccion__nivel_educativo__nivel'], estudiante['seccion__seccion']])
            ws.append([""])
            ws.append(["#", "Responsable", "DUI de responsable", "Estudiante", "Sección"])
            seccion = [estudiante['seccion__nivel_educativo__nivel'], estudiante['seccion__seccion']]
            contador_seccion = 1
        if estudiante["responsable__id"] in responsables_usados:
            count += 1
        else:
            ws.append(
                [
                    f"{contador_seccion}",
                    f"{estudiante['responsable__apellidos']}, {estudiante['responsable__nombre']}",
                    estudiante["responsable__dui"],
                    f"{estudiante['apellidos']}, {estudiante['nombre']}",
                    f"{estudiante['seccion__nivel_educativo__nivel']} {estudiante['seccion__seccion']}"
                ]
            )
            contador_seccion += 1
            for estudiante_b in estudiantes[count + 1 :]:
                if estudiante["responsable__id"] == estudiante_b["responsable__id"]:
                    ws.append(
                        [
                            "",
                            f"{estudiante_b['responsable__apellidos']}, {estudiante_b['responsable__nombre']}",
                            estudiante_b["responsable__dui"],
                            f"{estudiante_b['apellidos']}, {estudiante_b['nombre']}",
                            f"{estudiante_b['seccion__nivel_educativo__nivel']} {estudiante_b['seccion__seccion']}"
                        ]
                    )
            responsables_usados.append(estudiante["responsable__id"])
            count += 1

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/ms-excel",
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename=ListaDeEstudiantesYResponsables-{datetime.datetime.now().strftime("%Y_%m_%d-%H%M")}.xlsx'
        return response
    

def exportar_datos_de_contacto_tallaje_por_seccion(self, request, queryset):
    """
    Acción de Django Admin que exporta a Excel datos de contacto del modelo
    Estudiante haciendo un libro de excel por cada seccion, 
    de todos los estudiantes en el queryset. Para exportar todos lo libros estos
    se guardan en un zip y este zip es el respons
    Esto se hace haciendo uso de     la librería openpyxl ver: openpyxl.readthedocs.io
    """
    import zipfile
    from django.utils.text import slugify
    from tempfile import TemporaryDirectory

    with TemporaryDirectory() as temp_dir:
        zip_filename = f"DatosDeContactoPorSeccion-{datetime.datetime.now().strftime('%Y_%m_%d-%H%M')}.zip"
        zip_filepath = f"{temp_dir}/{zip_filename}"

        with zipfile.ZipFile(zip_filepath, 'w') as zipf:
            # Obtener las secciones únicas directamente desde el queryset
            secciones_ids = set(queryset.values_list('seccion__id', flat=True))
            
            for seccion_id in secciones_ids:
                # Filtrar estudiantes por sección
                estudiantes_seccion = queryset.filter(seccion__id=seccion_id).select_related("responsable", "seccion")
                
                # Verificar si hay estudiantes en esta sección
                if not estudiantes_seccion.exists():
                    continue
                
                wb = Workbook()
                ws = wb.active
                ws.column_dimensions['A'].width = 23
                ws.column_dimensions['B'].width = 23
                ws.column_dimensions['C'].width = 16
                ws.column_dimensions['D'].width = 11
                ws.column_dimensions['E'].width = 11
                ws.column_dimensions['K'].width = 16
                ws.column_dimensions['L'].width = 16 

                ws.merge_cells('A1:N1')
                ws['A1'] = 'Complejo Educativo "Dr. Humberto Romero Alvergue"'
                ws['A1'].alignment = ws['A1'].alignment.copy(horizontal='center', vertical='center')
                ws.title = "Estudiantes - contactos"
                ws.merge_cells('A2:N2')
                ws['A2'] = 'Tallaje de uniformes y calzado de estudiantes'
                ws['A2'].alignment = ws['A2'].alignment.copy(horizontal='center', vertical='center')
                ws['A2'].font = ws['A2'].font.copy(size=16)
                ws.merge_cells('A3:N3')
                ws['A3'] = f'Sección: {estudiantes_seccion[0].seccion.__str__().title()}'
                ws['A3'].alignment = ws['A3'].alignment.copy(horizontal='center', vertical='center')
                ws.append([])
                ws.merge_cells('A4:N4')
                ws['A4'] = 'Indicación: A continuación se le proporciona el detalle de contactos de los estudiantes de la sección, para que genere un grupo de WhatsApp de padres de familia, y hacer la consulta de tallas de calzado y uniforme del estudiante, de acuerdo con la clasificación proporcionada. Luego cargar los datos recolectados en el formulario enviado.'
                ws['A4'].alignment = ws['A4'].alignment.copy(horizontal='center', vertical='center')
                ws['A4'].alignment = ws['A4'].alignment.copy(wrap_text=True)
                ws.append([])

                ws.append(
                    [
                        "Apellidos",
                        "Nombres",
                        "NIE",
                        "Sexo",
                        "Edad",
                        "Teléfono de estudiante 1",
                        "Teléfono de estudiante 2",
                        "Nombre de responsable",
                        "Relación",
                        "DUI",
                        "Teléfono de responsable 1",
                        "Teléfono de responsable 2",
                        "Camisa (Talla)", # esto quedará en blanco para que lo llenen
                        "Pantalón/Falda (Talla)", # esto quedará en blanco para que lo llenen
                        "Calzado (Talla)", # esto quedará en blanco para que lo llenen
                    ]
                )
                for col in range(1, 16):  # Columnas A a O (1 a 15)
                    cell = ws.cell(row=6, column=col)
                    cell.font = cell.font.copy(bold=True)
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                    thin_border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    cell.border = thin_border
                # Obtener la sección del primer estudiante
                seccion_nombre = None
                for obj in estudiantes_seccion:
                    if seccion_nombre is None:
                        seccion_nombre = obj.seccion.__str__()
                    
                    responsable = obj.responsable
                    if responsable is None:
                        ws.append(
                            [
                                obj.apellidos,
                                obj.nombre,
                                obj.nie,
                                obj.sexo,
                                obj.edad,
                                obj.telefono_1,
                                obj.telefono_2,
                            ]
                        )
                    else:
                        ws.append(
                            [
                                obj.apellidos,
                                obj.nombre,
                                obj.nie,
                                obj.sexo,
                                obj.edad,
                                obj.telefono_1,
                                obj.telefono_2,
                                f"{responsable.nombre} {responsable.apellidos}",
                                obj.get_relacion_de_responsable_display(),
                                responsable.dui,
                                responsable.telefono_1,
                                responsable.telefono_2,
                            ]
                        )
                    current_row = ws.max_row
                    for col in range(1, 16): 
                        cell = ws.cell(row=current_row, column=col)
                        thin_border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.border = thin_border

                excel_filename = f"DatosDeContacto_{slugify(seccion_nombre)}.xlsx"
                excel_filepath = f"{temp_dir}/{excel_filename}"
                wb.save(excel_filepath)
                zipf.write(excel_filepath, arcname=excel_filename)
        with open(zip_filepath, 'rb') as zip_file:
            response = HttpResponse(
                zip_file.read(),
                content_type='application/zip',
            )
            response['Content-Disposition'] = f'attachment; filename={zip_filename}'
            return response

